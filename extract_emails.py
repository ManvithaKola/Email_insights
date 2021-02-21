import imaplib
import imapclient
import smtplib
import pprint
from collections import Counter
import pyzmail
import re
import urllib.parse
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook

imap_server = 'imap.gmail.com'
smtp_server = 'smtp.gmail.com'
username = 'email_ID'
password = 'password'

#Connect to servers and create access
imapobj = imapclient.IMAPClient(imap_server, ssl=True)#Enable Secure socket layer encryption
imapobj.login(username, password)

smtpobj = smtplib.SMTP(smtp_server, 587)
smtpobj.ehlo()
smtpobj.starttls() #Enable transport layer security TLS encryption
smtpobj.login(username, password)

pprint.pprint(imapobj.list_folders()) #Print the list of all folders in gmail

imaplib._MAXLINE = 10000000 #Increase memory allocation

imapobj.select_folder('Inbox', readonly=True) #Selecting only inbox folder for analysis
UIDs = imapobj.search(['SINCE', '01-Feb-2020', 'BEFORE', '20-Feb-2021']) #get the list of unique ID's for all emails

#Categorizizng the emails using get_gmail_labels
category = []
for i in range(len(UIDs)):
    label_dict = imapobj.get_gmail_labels(UIDs[i])
    label = label_dict[UIDs[i]]
    if 'Starred' in str(label):
        category.append('Starred')
    elif 'Important' in str(label):
        category.append('Important')
    elif len(label) == 0:
        category.append('Inbox')
    else:
        category.append('Custom Label')
print(Counter(category))

#Lists to hold different attributes of the mail
from_addresses = []
subjects = []
dates = []
days = []
months = []
years = []
times = []
sent_received = []
#make unsubscribing easy and a one-click process
#only unsubscribe links that have an address to which the unsubscribe
#email needs to be sent to and not the ones with an http link that needs to be clicked to unsubscribe
unsub_links = []

for i in range(len(UIDs)):
    raw_message = imapobj.fetch(UIDs[i], ['BODY[]'])
    message = pyzmail.PyzMessage.factory(raw_message[UIDs[i]][b'BODY[]'])

    if message.get_address('from')[1] == username:
        sent_received.append('Sent')
    else:
        sent_received.append('Received')

    full_date = message.get_decoded_header('date')
    from_addresses.append(message.get_address('from'))
    subjects.append(message.get_subject(''))
    unsub_link = message.get_decoded_header('List-Unsubscribe')
    if len(str(unsub_link)) > 0 and 'mailto' in unsub_link:
        unsub_link = unsub_link.split(',')
        unsub_links.append([unsub_link[idx] for idx, s in enumerate(unsub_link) if 'mailto' in s][0])
    else:
        unsub_links.append('No unsubscribe link found')

    day = full_date.split()[0].strip(',')
    date = full_date.split()[1]
    month = full_date.split()[2]
    year = full_date.split()[3]
    time = full_date.split()[4]
    days.append(day)
    dates.append(date)
    months.append(month)
    years.append(year)
    times.append(time)
    


imapobj.select_folder('Inbox', readonly=False)

del_UIDs = imapobj.search(['SUBJECT', 'The Corona Letter'])
print(del_UIDs)   
imapobj.delete_messages(del_UIDs)
imapobj.expunge()

# for Gmail, use this instead
imapobj.add_gmail_labels(del_UIDs,'\Trash')


wb = Workbook()
ws = wb.active
ws.title = "email_info"
ws.cell(1,1).value = "Date"
ws.cell(1,2).value = "Month"
ws.cell(1,3).value = "Year"
ws.cell(1,4).value = "Day"
ws.cell(1,5).value = "Time"
ws.cell(1,6).value = "From (Sender)"
ws.cell(1,7).value = "From (Email ID)"
ws.cell(1,8).value = "Subject"
ws.cell(1,9).value = "Sent/Received"
ws.cell(1,10).value = "Category"
for i in range(len(UIDs)):
    ws.cell(row=i+2, column=1).value = dates[i]
    ws.cell(row=i+2, column=2).value = months[i]
    ws.cell(row=i+2, column=3).value = years[i]
    ws.cell(row=i+2, column=4).value = days[i]
    ws.cell(row=i+2, column=5).value = times[i]
    ws.cell(row = i+2, column = 6).value = from_addresses[i][0]
    ws.cell(row = i+2, column = 7).value = from_addresses[i][1]
    ws.cell(row = i+2, column = 8).value = str(subjects[i])
    ws.cell(row=i + 2, column=9).value = sent_received[i]
    ws.cell(row=i+2, column=10).value = category[i]

wb.save('Email_Analytics.xlsx')

# disconnect from servers
smtpobj.quit()
imapobj.logout()
